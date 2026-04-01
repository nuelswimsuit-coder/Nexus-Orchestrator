"""
Management Ahu — Bridge API Router
Reads directly from the external Management Ahu (Telefix) project without copying its code.

Endpoints:
  GET  /api/ahu/status           — bot process running/stopped
  POST /api/ahu/bot/start        — launch run_bot.py
  POST /api/ahu/bot/stop         — kill bot process
  GET  /api/ahu/sessions         — session counts per folder (dynamic subdirs)
  POST /api/ahu/sessions/sync-scanned — copy new sessions from scanner → כללי
  POST /api/ahu/sessions/move    — move session files between folders
  GET  /api/ahu/stats            — DB stats; users also union JSON/CSV under data/קהיל חיה (TELEFIX_DISK_USERS_DIR)
  GET  /api/ahu/targets          — source + target groups from DB
  GET  /api/ahu/logs             — last N lines from telefix.log
  WS   /api/ahu/logs/stream      — live log streaming via WebSocket
"""

from __future__ import annotations

import asyncio
import csv
import json
import os
import shutil
import sqlite3
import subprocess
import sys
from pathlib import Path
from typing import Any

import structlog
from fastapi import APIRouter, Depends, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field

from nexus.shared.config import settings

log = structlog.get_logger(__name__)


async def _require_legacy_telefix_bot() -> None:
    if not settings.legacy_telefix_bot_enabled:
        raise HTTPException(
            status_code=503,
            detail="Legacy TeleFix bot (AHU) is disabled. Use the management dashboard API at /api/management.",
        )


router = APIRouter(
    prefix="/ahu",
    tags=["ahu"],
    dependencies=[Depends(_require_legacy_telefix_bot)],
)

# Default folder for merged / imported operator-scanned sessions (Hebrew name on disk)
KLALI_FOLDER = "כללי"

# Nexus repo root (…/Nexus-Orchestrator) — for default NEXUS_OPERATOR_SCAN_DIR
_NEXUS_REPO_ROOT = Path(__file__).resolve().parents[3]

# Global reference to the bot subprocess (one at a time)
_bot_process: subprocess.Popen | None = None


# ── Paths (TELEFIX_ROOT / TELEFIX_SESSIONS_DIR / TELEFIX_DB) ───────────────────


def _telefix_root() -> Path:
    raw = os.environ.get("TELEFIX_ROOT", "").strip() or settings.telefix_root
    return Path(raw).expanduser().resolve()


def _ahu_sessions_dir() -> Path:
    raw = (os.environ.get("TELEFIX_SESSIONS_DIR") or settings.telefix_sessions_dir or "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    return _telefix_root() / "sessions"


def _ahu_db_path() -> Path:
    raw = (os.environ.get("TELEFIX_DB") or settings.telefix_db or "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    return _telefix_root() / "data" / "telefix.db"


def _ahu_log_path() -> Path:
    return _telefix_root() / "logs" / "telefix.log"


def _ahu_bot_path() -> Path:
    return _telefix_root() / "run_bot.py"


def _operator_scan_dir() -> Path:
    raw = os.environ.get("NEXUS_OPERATOR_SCAN_DIR", "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    return _NEXUS_REPO_ROOT / "sessions" / "validated_active"


def _disk_users_dir() -> Path:
    """
    Extra scraped users on disk (e.g. exports under Telefix ``data/קהיל חיה``).
    Override with TELEFIX_DISK_USERS_DIR.
    """
    raw = os.environ.get("TELEFIX_DISK_USERS_DIR", "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    return _telefix_root() / "data" / "קהיל חיה"


def _parse_int_user_id(val: Any) -> int | None:
    try:
        if val is None or val == "":
            return None
        return int(val)
    except (TypeError, ValueError):
        return None


def _record_premium_flag(rec: dict[str, Any]) -> int:
    v = rec.get("is_premium")
    if v is None:
        v = rec.get("isPremium") or rec.get("premium")
    if v in (True, 1, "1", "true", "True"):
        return 1
    return 0


def _record_user_id(rec: dict[str, Any]) -> int | None:
    for key in ("user_id", "userId", "id", "telegram_id", "telegramId"):
        if key in rec:
            uid = _parse_int_user_id(rec.get(key))
            if uid is not None:
                return uid
    return None


def _normalize_phone_cell(val: Any) -> int | None:
    """Excel / CSV phone → int digits (Account Summary «Phone» column)."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, float):
        if val != val or abs(val) > 1e15:  # nan or huge
            return None
        val = int(round(val))
    s = str(val).strip().replace("+", "").replace("-", "").replace(" ", "")
    if not s or not s.isdigit():
        return None
    if len(s) < 9:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _record_source_group(rec: dict[str, Any]) -> str:
    s = (
        rec.get("source_group")
        or rec.get("sourceGroup")
        or rec.get("origin_group")
        or ""
    )
    t = str(s).strip()
    return t if t else "disk:קהיל חיה"


def _iter_disk_scraped_records(root: Path) -> list[dict[str, Any]]:
    """Load user dicts from JSON bundles, CSV, or per-file JSON under root."""
    out: list[dict[str, Any]] = []
    if not root.is_dir():
        return out

    for name in ("scraped_users.json", "users.json", "scraped.json"):
        p = root / name
        if not p.is_file():
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8", errors="replace"))
        except Exception:
            continue
        if isinstance(data, list):
            out.extend([x for x in data if isinstance(x, dict)])
        elif isinstance(data, dict):
            inner = data.get("users") or data.get("scraped_users")
            if isinstance(inner, list):
                out.extend([x for x in inner if isinstance(x, dict)])

    for p in sorted(root.glob("*.csv")):
        try:
            with p.open(encoding="utf-8", errors="replace", newline="") as f:
                for row in csv.DictReader(f):
                    if isinstance(row, dict):
                        out.append(dict(row))
        except Exception:
            pass

    bundled = {"scraped_users.json", "users.json", "scraped.json"}
    for p in sorted(root.glob("*.json")):
        if p.name in bundled:
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8", errors="replace"))
        except Exception:
            continue
        if isinstance(data, dict):
            out.append(data)
        elif isinstance(data, list):
            out.extend([x for x in data if isinstance(x, dict)])

    return out


def _iter_audit_xlsx_account_summary(repo_root: Path) -> list[dict[str, Any]]:
    """
    Rows from ``nexus_audit_*.xlsx`` → sheet ``Account Summary`` (Phone per account).

    These are Telethon accounts audited on disk — not the same rows as
    ``scraped_users`` in telefix.db (those are scraped *members*). openpyxl optional.
    """
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        return []

    out: list[dict[str, Any]] = []
    for path in sorted(repo_root.glob("nexus_audit_*.xlsx")):
        wb: Any = None
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if "Account Summary" not in wb.sheetnames:
                continue
            ws = wb["Account Summary"]
            rows = ws.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            h = [str(x).strip() if x is not None else "" for x in header]
            try:
                pi = h.index("Phone")
            except ValueError:
                continue
            prem_i: int | None = None
            for cand in ("Premium", "Premium Acc"):
                if cand in h:
                    prem_i = h.index(cand)
                    break
            for row in rows:
                if not row:
                    continue
                cell = row[pi] if pi < len(row) else None
                phone = _normalize_phone_cell(cell)
                if phone is None:
                    continue
                pr = 0
                if prem_i is not None and prem_i < len(row):
                    pv = row[prem_i]
                    if str(pv).strip().lower() in ("yes", "true", "1", "premium"):
                        pr = 1
                out.append(
                    {
                        "user_id": phone,
                        "is_premium": pr,
                        "source_group": f"audit:{path.name}",
                    }
                )
        except Exception:
            continue
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
    return out


def _iter_group_audit_csv_session_ids(repo_root: Path) -> list[dict[str, Any]]:
    """``nexus_group_audit.csv`` — numeric Session_Name → synthetic id (group audit export)."""
    p = repo_root / "nexus_group_audit.csv"
    if not p.is_file():
        return []
    out: list[dict[str, Any]] = []
    try:
        with p.open(encoding="utf-8", errors="replace", newline="") as f:
            for row in csv.DictReader(f):
                sn = (row.get("Session_Name") or "").strip()
                if not sn.isdigit() or len(sn) < 9:
                    continue
                uid = int(sn)
                out.append(
                    {
                        "user_id": uid,
                        "is_premium": 0,
                        "source_group": "audit:nexus_group_audit.csv",
                    }
                )
    except Exception:
        pass
    return out


def _merge_disk_scraped_users(
    *,
    db_total: int,
    db_premium: int,
    db_sources: int,
) -> tuple[int, int, int, int, str]:
    """
    Union SQLite scraped_users with:

    - JSON/CSV under ``data/קהיל חיה`` (TELEFIX_DISK_USERS_DIR)
    - ``nexus_audit_*.xlsx`` → ``Account Summary`` (Phone = account id)
    - ``nexus_group_audit.csv`` numeric Session_Name (session stem)

    Deduplicates by numeric id. DB rows are Telegram *scraped members*; audit
    files are *accounts/sessions* — different concepts, merged only for dashboard cardinality.

    Returns:
        (merged_total, merged_premium, merged_source_count, disk_only_count, disk_dir_str)
    """
    root = _disk_users_dir()
    disk_dir_str = str(root)

    id_rows = _query("SELECT user_id FROM scraped_users")
    db_ids: set[int] = set()
    for r in id_rows:
        try:
            db_ids.add(int(r["user_id"]))
        except (TypeError, ValueError, KeyError):
            pass

    src_rows = _query(
        "SELECT DISTINCT source_group FROM scraped_users "
        "WHERE source_group IS NOT NULL AND trim(source_group) != ''"
    )
    source_names: set[str] = set()
    for r in src_rows:
        try:
            sg = str(r["source_group"]).strip()
            if sg:
                source_names.add(sg)
        except Exception:
            pass

    records: list[dict[str, Any]] = []
    if root.is_dir():
        records.extend(_iter_disk_scraped_records(root))
    records.extend(_iter_audit_xlsx_account_summary(_NEXUS_REPO_ROOT))
    records.extend(_iter_group_audit_csv_session_ids(_NEXUS_REPO_ROOT))

    if not records:
        return db_total, db_premium, db_sources, 0, disk_dir_str

    by_id: dict[int, dict[str, Any]] = {}
    for rec in records:
        uid = _record_user_id(rec)
        if uid is None:
            continue
        by_id[uid] = rec

    disk_ids = set(by_id.keys())
    merged_ids = db_ids | disk_ids
    merged_total = len(merged_ids)

    disk_only = disk_ids - db_ids
    extra_premium = sum(_record_premium_flag(by_id[u]) for u in disk_only)
    merged_premium = db_premium + extra_premium

    for u in disk_only:
        source_names.add(_record_source_group(by_id[u]))

    merged_sources = len(source_names)
    disk_only_count = len(disk_only)

    return merged_total, merged_premium, merged_sources, disk_only_count, disk_dir_str


# ── Helpers ────────────────────────────────────────────────────────────────────


def _db_connect() -> sqlite3.Connection | None:
    """Return a read-only SQLite connection to telefix.db, or None if absent."""
    p = _ahu_db_path()
    if not p.exists():
        return None
    try:
        conn = sqlite3.connect(f"file:{p.as_posix()}?mode=ro", uri=True, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn
    except Exception as exc:
        log.warning("ahu_db_connect_failed", error=str(exc))
        return None


def _query(sql: str, params: tuple = ()) -> list[dict]:
    conn = _db_connect()
    if conn is None:
        return []
    try:
        cur = conn.execute(sql, params)
        rows = [dict(r) for r in cur.fetchall()]
        return rows
    except Exception as exc:
        log.warning("ahu_db_query_failed", sql=sql, error=str(exc))
        return []
    finally:
        conn.close()


def _query_one(sql: str, params: tuple = ()) -> dict | None:
    rows = _query(sql, params)
    return rows[0] if rows else None


def _ensure_klali(sessions_root: Path) -> None:
    (sessions_root / KLALI_FOLDER).mkdir(parents=True, exist_ok=True)


def _collect_stems_in_folder(cat_dir: Path) -> list[str]:
    sessions: list[str] = []
    if not cat_dir.is_dir():
        return sessions
    try:
        for item in cat_dir.iterdir():
            if item.suffix == ".session":
                sessions.append(item.stem)
            elif item.is_dir():
                inner = list(item.glob("*.session"))
                if inner:
                    sessions.append(item.name)
    except OSError:
        pass
    return sorted(sessions)


def _scan_sessions() -> dict[str, Any]:
    """Scan every immediate subdirectory under sessions/ for .session entries."""
    sessions_root = _ahu_sessions_dir()
    _ensure_klali(sessions_root)
    result: dict[str, Any] = {}
    if not sessions_root.exists():
        return result
    subdirs = sorted(
        [p for p in sessions_root.iterdir() if p.is_dir()],
        key=lambda p: p.name.lower(),
    )
    for cat_dir in subdirs:
        name = cat_dir.name
        stems = _collect_stems_in_folder(cat_dir)
        result[name] = {"count": len(stems), "sessions": stems}
    return result


def _unique_stem_count(data: dict[str, Any]) -> int:
    seen: set[str] = set()
    for v in data.values():
        if not isinstance(v, dict):
            continue
        for s in v.get("sessions") or []:
            seen.add(str(s))
    return len(seen)


def _all_stems_under_vault(sessions_root: Path) -> set[str]:
    _ensure_klali(sessions_root)
    out: set[str] = set()
    if not sessions_root.exists():
        return out
    for cat_dir in sessions_root.iterdir():
        if not cat_dir.is_dir():
            continue
        for s in _collect_stems_in_folder(cat_dir):
            out.add(s)
    return out


def _safe_subfolder(name: str) -> str:
    n = (name or "").strip()
    if not n or ".." in n or "/" in n or "\\" in n:
        raise HTTPException(status_code=400, detail="invalid folder name")
    return n


def _resolve_stem_location(folder: Path, stem: str) -> tuple[Path, str] | None:
    """
    Return (path_to_move, kind) where kind is 'file' or 'dir'.
    path_to_move is the .session file or the numbered subdirectory.
    """
    direct = folder / f"{stem}.session"
    if direct.is_file():
        return (direct, "file")
    sub = folder / stem
    if sub.is_dir() and list(sub.glob("*.session")):
        return (sub, "dir")
    return None


def _tail_log(n: int = 150) -> list[str]:
    """Return the last n lines from telefix.log."""
    p = _ahu_log_path()
    if not p.exists():
        return []
    try:
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        return [l.rstrip() for l in lines[-n:]]
    except Exception as exc:
        log.warning("ahu_log_read_failed", error=str(exc))
        return []


def _bot_is_running() -> bool:
    global _bot_process
    if _bot_process is None:
        return False
    poll = _bot_process.poll()
    if poll is not None:
        _bot_process = None
        return False
    return True


# ── Status ─────────────────────────────────────────────────────────────────────


@router.get("/status")
async def get_status() -> JSONResponse:
    running = _bot_is_running()
    root = _telefix_root()
    db_ok = _ahu_db_path().exists()
    sessions_root = _ahu_sessions_dir()
    sessions_ok = sessions_root.exists()

    session_counts: dict[str, int] = {}
    total_sessions = 0
    if sessions_ok:
        cats = _scan_sessions()
        session_counts = {k: v["count"] for k, v in cats.items()}
        total_sessions = _unique_stem_count(cats)

    return JSONResponse({
        "bot_running": running,
        "bot_pid": _bot_process.pid if running and _bot_process else None,
        "db_available": db_ok,
        "sessions_available": sessions_ok,
        "total_sessions": total_sessions,
        "session_counts": session_counts,
        "ahu_root": str(root),
    })


# ── Bot control ────────────────────────────────────────────────────────────────


@router.post("/bot/start")
async def start_bot() -> JSONResponse:
    global _bot_process
    bot = _ahu_bot_path()
    root = _telefix_root()
    if _bot_is_running():
        return JSONResponse({"ok": False, "detail": "Bot is already running", "pid": _bot_process.pid})
    if not bot.exists():
        return JSONResponse({"ok": False, "detail": f"run_bot.py not found at {bot}"}, status_code=404)
    try:
        _bot_process = subprocess.Popen(
            [sys.executable, str(bot)],
            cwd=str(root),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        log.info("ahu_bot_started", pid=_bot_process.pid)
        return JSONResponse({"ok": True, "pid": _bot_process.pid})
    except Exception as exc:
        log.error("ahu_bot_start_failed", error=str(exc))
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=500)


@router.post("/bot/stop")
async def stop_bot() -> JSONResponse:
    global _bot_process
    if not _bot_is_running():
        return JSONResponse({"ok": False, "detail": "Bot is not running"})
    try:
        pid = _bot_process.pid
        _bot_process.terminate()
        try:
            _bot_process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            _bot_process.kill()
        _bot_process = None
        log.info("ahu_bot_stopped", pid=pid)
        return JSONResponse({"ok": True, "pid": pid})
    except Exception as exc:
        log.error("ahu_bot_stop_failed", error=str(exc))
        return JSONResponse({"ok": False, "detail": str(exc)}, status_code=500)


# ── Sessions ───────────────────────────────────────────────────────────────────


@router.get("/sessions")
async def get_sessions() -> JSONResponse:
    data = _scan_sessions()
    return JSONResponse(data)


class MoveSessionBody(BaseModel):
    stem: str = Field(..., min_length=1)
    from_folder: str = Field(..., min_length=1)
    to_folder: str = Field(..., min_length=1)


@router.post("/sessions/move")
async def move_session(body: MoveSessionBody) -> JSONResponse:
    if body.from_folder == body.to_folder:
        raise HTTPException(status_code=400, detail="from_folder and to_folder must differ")
    src_name = _safe_subfolder(body.from_folder)
    dst_name = _safe_subfolder(body.to_folder)
    stem = body.stem.strip()
    if not stem:
        raise HTTPException(status_code=400, detail="invalid stem")

    sessions_root = _ahu_sessions_dir()
    _ensure_klali(sessions_root)
    src_dir = sessions_root / src_name
    dst_dir = sessions_root / dst_name
    if not src_dir.is_dir():
        raise HTTPException(status_code=404, detail="source folder not found")
    dst_dir.mkdir(parents=True, exist_ok=True)

    loc = _resolve_stem_location(src_dir, stem)
    if loc is None:
        raise HTTPException(status_code=404, detail="session not found in source folder")

    path_obj, kind = loc

    # Collision check in destination
    if _resolve_stem_location(dst_dir, stem) is not None:
        raise HTTPException(status_code=409, detail="destination already has this session")

    try:
        if kind == "file":
            dest_file = dst_dir / path_obj.name
            shutil.move(str(path_obj), str(dest_file))
            journal = src_dir / f"{stem}.session-journal"
            if journal.exists():
                shutil.move(str(journal), str(dst_dir / journal.name))
            meta = src_dir / f"{stem}.json"
            if meta.exists():
                shutil.move(str(meta), str(dst_dir / meta.name))
        else:
            dest_sub = dst_dir / path_obj.name
            shutil.move(str(path_obj), str(dest_sub))
    except HTTPException:
        raise
    except Exception as exc:
        log.warning("ahu_session_move_failed", error=str(exc))
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    return JSONResponse({"ok": True, "stem": stem, "from_folder": src_name, "to_folder": dst_name})


@router.post("/sessions/sync-scanned")
async def sync_scanned_sessions() -> JSONResponse:
    """
    Copy *.session (+ .json / .session-journal) from the operator scan output
    into sessions/כללי when the stem is not already present anywhere under sessions/.
    """
    sessions_root = _ahu_sessions_dir()
    _ensure_klali(sessions_root)
    dest = sessions_root / KLALI_FOLDER
    dest.mkdir(parents=True, exist_ok=True)

    src_root = _operator_scan_dir()
    if not src_root.is_dir():
        return JSONResponse(
            {"ok": False, "detail": f"scan directory not found: {src_root}", "copied": 0, "skipped": 0}
        )

    existing = _all_stems_under_vault(sessions_root)
    copied = 0
    skipped = 0
    errors: list[str] = []

    for sess_file in sorted(src_root.glob("*.session")):
        stem = sess_file.stem
        if stem in existing:
            skipped += 1
            continue
        try:
            shutil.copy2(str(sess_file), str(dest / sess_file.name))
            existing.add(stem)
            copied += 1
            js = sess_file.with_suffix(".json")
            if js.is_file():
                shutil.copy2(str(js), str(dest / js.name))
            jrn = Path(str(sess_file) + "-journal")
            if jrn.is_file():
                shutil.copy2(str(jrn), str(dest / jrn.name))
        except Exception as exc:
            errors.append(f"{stem}: {exc}")
            log.warning("ahu_sync_copy_failed", stem=stem, error=str(exc))

    return JSONResponse(
        {
            "ok": True,
            "copied": copied,
            "skipped": skipped,
            "source": str(src_root),
            "destination": str(dest),
            "errors": errors[:20],
        }
    )


# ── Stats ──────────────────────────────────────────────────────────────────────


@router.get("/stats")
async def get_stats() -> JSONResponse:
    # scraped_users (SQLite) + optional union with JSON/CSV under data/קהיל חיה
    users_row = _query_one("SELECT COUNT(*) as total, SUM(is_premium) as premium FROM scraped_users")
    total_users = int(users_row["total"]) if users_row else 0
    premium_users = int(users_row["premium"] or 0) if users_row else 0

    sources_row = _query_one("SELECT COUNT(DISTINCT source_group) as cnt FROM scraped_users")
    total_sources = int(sources_row["cnt"]) if sources_row else 0

    merged_total, merged_premium, merged_sources, disk_only_count, disk_users_dir = _merge_disk_scraped_users(
        db_total=total_users,
        db_premium=premium_users,
        db_sources=total_sources,
    )

    # targets
    targets_rows = _query("SELECT role, COUNT(*) as cnt FROM targets GROUP BY role")
    targets_by_role: dict[str, int] = {}
    for r in targets_rows:
        targets_by_role[r["role"]] = int(r["cnt"])

    # enrollments
    enroll_row = _query_one("SELECT COUNT(*) as total FROM enrollments")
    total_enrollments = int(enroll_row["total"]) if enroll_row else 0

    enroll_by_status = _query("SELECT status, COUNT(*) as cnt FROM enrollments GROUP BY status")
    enrollment_status: dict[str, int] = {r["status"]: int(r["cnt"]) for r in enroll_by_status}

    # last run metrics
    metrics = _query("SELECT key, value FROM settings WHERE key LIKE 'last_run:%'")
    last_runs: dict[str, str] = {r["key"].replace("last_run:", ""): r["value"] for r in metrics}

    return JSONResponse({
        "users": {
            "total": merged_total,
            "premium": merged_premium,
            "sources": merged_sources,
            "premium_pct": round(merged_premium * 100 / merged_total, 1) if merged_total > 0 else 0,
            "disk_only_count": disk_only_count,
            "disk_users_dir": disk_users_dir,
        },
        "targets": targets_by_role,
        "enrollments": {
            "total": total_enrollments,
            "by_status": enrollment_status,
        },
        "last_runs": last_runs,
    })


# ── Targets ────────────────────────────────────────────────────────────────────


@router.get("/targets")
async def get_targets() -> JSONResponse:
    rows = _query("SELECT id, title, link, role FROM targets ORDER BY role, id")
    return JSONResponse({"targets": rows, "count": len(rows)})


# ── Scraped users (paginated) ──────────────────────────────────────────────────


@router.get("/scraped-users")
async def get_scraped_users(limit: int = 50, offset: int = 0) -> JSONResponse:
    rows = _query(
        "SELECT user_id, username, source_group, is_premium, added_at "
        "FROM scraped_users ORDER BY added_at DESC LIMIT ? OFFSET ?",
        (limit, offset),
    )
    total_row = _query_one("SELECT COUNT(*) as cnt FROM scraped_users")
    total = int(total_row["cnt"]) if total_row else 0
    return JSONResponse({"users": rows, "total": total, "limit": limit, "offset": offset})


# ── Logs ───────────────────────────────────────────────────────────────────────


@router.get("/logs")
async def get_logs(lines: int = 150) -> JSONResponse:
    log_lines = _tail_log(min(lines, 500))
    return JSONResponse({"lines": log_lines, "count": len(log_lines)})


@router.websocket("/logs/stream")
async def stream_logs(ws: WebSocket) -> None:
    await ws.accept()
    log_path = _ahu_log_path()
    if not log_path.exists():
        await ws.send_text("[ahu] telefix.log not found")
        await ws.close()
        return

    # Send last 50 lines as initial burst
    initial = _tail_log(50)
    for line in initial:
        await ws.send_text(line)

    # Then tail new lines
    try:
        with open(log_path, "r", encoding="utf-8", errors="replace") as f:
            f.seek(0, 2)  # seek to end
            while True:
                line = f.readline()
                if line:
                    await ws.send_text(line.rstrip())
                else:
                    await asyncio.sleep(0.5)
    except WebSocketDisconnect:
        pass
    except Exception as exc:
        log.warning("ahu_log_stream_error", error=str(exc))
        try:
            await ws.close()
        except Exception:
            pass
