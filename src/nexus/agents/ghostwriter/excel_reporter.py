"""
Excel Reporter — generates a formatted .xlsx report from audit results.
Sorted by member_count DESC per entity, with color-coded status rows.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side
)
from openpyxl.utils import get_column_letter

from .session_auditor import AccountAudit, EntityAudit


# ── Colour palette ────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_OWNER_BG    = "D6E4F0"   # light blue  — owner rows
C_ADMIN_BG    = "EBF5FB"   # lighter blue — admin rows
C_BANNED_BG   = "FADBD8"   # red tint
C_DEAD_BG     = "F2F3F4"   # grey
C_PREMIUM_FG  = "7D3C98"   # purple for premium counts
C_WORKING_BG  = "EAFAF1"   # green tint — working accounts
C_ALT_ROW     = "F8F9FA"   # alternating row

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_font() -> Font:
    return Font(bold=True, color=C_HEADER_FG, size=11)


def _bold(color: str = "000000") -> Font:
    return Font(bold=True, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 1: Entities (groups/channels) sorted by member count ───────────────

def _build_entities_sheet(wb: openpyxl.Workbook, audits: list[AccountAudit]) -> None:
    ws = wb.active
    ws.title = "Groups & Channels"

    headers = [
        "Phone", "Username", "Name", "Premium Acc",
        "Entity Title", "Type", "Role", "Link",
        "Members", "Premium (Real)", "Premium (Boosts)",
        "Premium Real %", "Session Status",
    ]

    # Header row
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill   = _fill(C_HEADER_BG)
        cell.font   = _header_font()
        cell.alignment = _center()
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Flatten all entity rows
    rows: list[tuple] = []
    for audit in audits:
        status = (
            "banned"        if audit.is_banned else
            "unregistered"  if audit.is_unregistered else
            "error"         if audit.error else
            "working"
        )
        for ent in audit.entities:
            pct = round(ent.premium_real / ent.member_count * 100, 1) if ent.member_count else 0
            rows.append((
                audit.phone,
                audit.username,
                f"{audit.first_name} {audit.last_name}".strip(),
                "YES" if audit.is_premium else "no",
                ent.title,
                ent.entity_type,
                ent.role,
                ent.invite_link,
                ent.member_count,
                ent.premium_real,
                ent.premium_boosts,
                pct,
                status,
            ))

    # Sort by member_count DESC
    rows.sort(key=lambda r: r[8], reverse=True)

    for row_idx, row_data in enumerate(rows, 2):
        ws.append(list(row_data))
        role   = row_data[6]
        status = row_data[12]

        bg = (
            C_BANNED_BG if status == "banned" else
            C_DEAD_BG   if status in ("unregistered", "error") else
            C_OWNER_BG  if role == "owner" else
            C_ADMIN_BG
        )

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill      = _fill(bg)
            cell.border    = BORDER
            cell.alignment = _center() if col in (1, 6, 7, 9, 10, 11, 12) else _left()

        # Purple for premium columns
        for col in (10, 11, 12):
            ws.cell(row=row_idx, column=col).font = _bold(C_PREMIUM_FG)

        # Bold member count
        ws.cell(row=row_idx, column=9).font = _bold()

    _set_col_widths(ws, [18, 18, 22, 12, 35, 10, 8, 35, 12, 15, 15, 14, 14])


# ── Sheet 2: Account summary ──────────────────────────────────────────────────

def _build_accounts_sheet(wb: openpyxl.Workbook, audits: list[AccountAudit]) -> None:
    ws = wb.create_sheet("Account Summary")

    headers = [
        "Phone", "Username", "Name", "Premium",
        "Status", "Groups Owned", "Channels Owned",
        "Groups Admin", "Channels Admin",
        "Total Entities", "Error",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill      = _fill(C_HEADER_BG)
        cell.font      = _header_font()
        cell.alignment = _center()
        cell.border    = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for row_idx, audit in enumerate(audits, 2):
        status = (
            "BANNED"        if audit.is_banned else
            "DEAD"          if audit.is_unregistered else
            "ERROR"         if audit.error else
            "WORKING"
        )
        groups_owned    = sum(1 for e in audit.entities if e.entity_type == "group"   and e.role == "owner")
        channels_owned  = sum(1 for e in audit.entities if e.entity_type == "channel" and e.role == "owner")
        groups_admin    = sum(1 for e in audit.entities if e.entity_type == "group"   and e.role == "admin")
        channels_admin  = sum(1 for e in audit.entities if e.entity_type == "channel" and e.role == "admin")

        ws.append([
            audit.phone,
            audit.username,
            f"{audit.first_name} {audit.last_name}".strip(),
            "YES" if audit.is_premium else "no",
            status,
            groups_owned, channels_owned,
            groups_admin, channels_admin,
            len(audit.entities),
            audit.error,
        ])

        bg = (
            C_BANNED_BG  if audit.is_banned else
            C_DEAD_BG    if audit.is_unregistered else
            C_DEAD_BG    if audit.error else
            C_WORKING_BG
        )
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill      = _fill(bg if row_idx % 2 == 0 else bg)
            cell.border    = BORDER
            cell.alignment = _center()

    _set_col_widths(ws, [18, 18, 22, 10, 12, 14, 16, 14, 16, 14, 30])


# ── Sheet 3: Stats summary ────────────────────────────────────────────────────

def _build_stats_sheet(wb: openpyxl.Workbook, audits: list[AccountAudit]) -> None:
    ws = wb.create_sheet("Stats")

    total      = len(audits)
    working    = sum(1 for a in audits if not a.error and not a.is_banned and not a.is_unregistered)
    banned     = sum(1 for a in audits if a.is_banned)
    dead       = sum(1 for a in audits if a.is_unregistered)
    errors     = sum(1 for a in audits if a.error and not a.is_banned and not a.is_unregistered)
    premium_acc = sum(1 for a in audits if a.is_premium)

    all_entities = [e for a in audits for e in a.entities]
    total_groups   = sum(1 for e in all_entities if e.entity_type == "group")
    total_channels = sum(1 for e in all_entities if e.entity_type == "channel")
    total_members  = sum(e.member_count for e in all_entities)
    total_premium_real   = sum(e.premium_real for e in all_entities)
    total_premium_boosts = sum(e.premium_boosts for e in all_entities)

    stats = [
        ("Scan Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("ACCOUNTS", ""),
        ("Total Sessions Audited", total),
        ("Working", working),
        ("Banned", banned),
        ("Dead / Unregistered", dead),
        ("Errors", errors),
        ("Premium Accounts", premium_acc),
        ("", ""),
        ("ENTITIES (owned/admin)", ""),
        ("Total Groups", total_groups),
        ("Total Channels", total_channels),
        ("Total Members (all entities)", total_members),
        ("Total Premium Members (real scan)", total_premium_real),
        ("Total Premium Members (boosts panel)", total_premium_boosts),
    ]

    for row_idx, (label, value) in enumerate(stats, 1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if label and not label.startswith("  "):
            ws.cell(row=row_idx, column=1).font = _bold()
        if isinstance(value, int):
            ws.cell(row=row_idx, column=2).font = _bold(C_PREMIUM_FG)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 25


# ── Public API ────────────────────────────────────────────────────────────────

def generate_report(audits: list[AccountAudit], output_path: Path | None = None) -> Path:
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = Path(f"nexus_audit_{ts}.xlsx")

    wb = openpyxl.Workbook()
    _build_entities_sheet(wb, audits)
    _build_accounts_sheet(wb, audits)
    _build_stats_sheet(wb, audits)

    wb.save(output_path)
    return output_path
